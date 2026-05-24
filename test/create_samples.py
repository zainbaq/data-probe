from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

header_font = Font(name="Arial", bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="2E4057")
cell_font   = Font(name="Arial", size=10)

def style_headers(ws, count):
    for col in range(1, count + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

def style_data_rows(ws, headers):
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).font = cell_font

# ── sales_data.xlsx ───────────────────────────────────────────────────────────
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Sales Data"

headers1 = ["order_id","customer_name","email","product","category",
            "quantity","unit_price","total_price","order_date","region","status"]
ws1.append(headers1)
style_headers(ws1, len(headers1))

sales_rows = [
    ["ORD-001","Alice Johnson","alice@example.com","Laptop Pro","Electronics",2,1299.99,"=G2*F2","2024-01-05","North","Completed"],
    ["ORD-002","Bob Smith","bob@example.com","Office Chair","Furniture",4,249.99,"=G3*F3","2024-01-07","South","Completed"],
    ["ORD-003","Carol White","carol@example.com","Wireless Mouse","Electronics",10,29.99,"=G4*F4","2024-01-09","East","Completed"],
    ["ORD-004","David Lee","david@example.com","Standing Desk","Furniture",1,899.99,"=G5*F5","2024-01-11","West","Pending"],
    ["ORD-005","Eve Davis","eve@example.com","Monitor 27in","Electronics",3,399.99,"=G6*F6","2024-01-12","North","Completed"],
    ["ORD-006","Frank Brown","frank_at_example.com","Keyboard","Electronics",5,79.99,"=G7*F7","2024-01-14","East","Completed"],
    ["ORD-007","Grace Kim","grace@example.com","Webcam","Electronics",2,149.99,"=G8*F8","2024-01-15","South","Returned"],
    ["ORD-008","Henry Clark","henry@example.com","Notebook","Stationery",50,4.99,"=G9*F9","2024-01-16","West","Completed"],
    ["ORD-009","Iris Lopez","iris@example.com","Laptop Pro","Electronics",1,1299.99,"=G10*F10","2024-01-18","North","Completed"],
    ["ORD-010","Jack Patel","jack@example.com","USB Hub","Electronics",8,39.99,"=G11*F11","Jan 20 2024","East","Completed"],
    ["ORD-011","Karen Ng","karen@example.com","Office Chair","Furniture",2,249.99,"=G12*F12","2024-01-21","South","Completed"],
    ["ORD-012","Leo Turner","leo@example.com","Printer","Electronics",1,599.99,"=G13*F13","2024-01-22","North","Pending"],
    ["ORD-013","Mia Scott","mia@example.com","Desk Lamp","Furniture",6,34.99,"=G14*F14","2024-01-23","West","Completed"],
    ["ORD-014","Nate Hill","nate@example.com","Webcam","Electronics",1,149.99,"=G15*F15","2024-01-24","East","Completed"],
    ["ORD-015","Olivia Reed","olivia@example.com","Laptop Pro","Electronics",1,99999.99,"=G16*F16","2024-01-25","South","Completed"],
    ["ORD-016","Paul Adams","paul@example.com","Mouse Pad","Stationery",20,9.99,"=G17*F17","2024-01-26","North","Completed"],
    ["ORD-017","Quinn Ross","quinn@example.com","Monitor 27in","Electronics",2,399.99,"=G18*F18","2024-01-27","East","Pending"],
    ["ORD-018","Rachel Tran","rachel@example.com","Standing Desk","Furniture",1,899.99,"=G19*F19","2024-01-28","West","Completed"],
    ["ORD-019","Sam Wright","sam@example.com","Keyboard","Electronics",3,79.99,"=G20*F20","2024-01-29","South","Completed"],
    ["ORD-020","Tina Evans","","USB Hub","Electronics",4,39.99,"=G21*F21","2024-01-30","North","Completed"],
    ["ORD-021","Uma Flores","uma@example.com","Notebook","Stationery",30,4.99,"=G22*F22","2024-01-31","East","Completed"],
    ["ORD-022","Victor Stone","victor@example.com","Laptop Pro","Electronics",2,1299.99,"=G23*F23","2024-02-01","West","Completed"],
    ["ORD-023","Wendy Cruz","wendy@example.com","Office Chair","Furniture",3,249.99,"=G24*F24","2024-02-02","South","Completed"],
    ["ORD-024","Xander Bell","xander@example.com","Wireless Mouse","Electronics",7,29.99,"=G25*F25","2024-02-03","North","Returned"],
    ["ORD-025","Yara Collins","yara@example.com","Monitor 27in","Electronics",1,399.99,"=G26*F26","2024-02-04","East","Completed"],
    ["ORD-026","Zoe Morris","zoe@example.com","Printer","Electronics",2,599.99,"=G27*F27","2024-02-05","West","Completed"],
    ["ORD-027","Alice Johnson","alice@example.com","Laptop Pro","Electronics",2,1299.99,"=G28*F28","2024-01-05","North","Completed"],
    ["ORD-028","Brian Hall","brian@example.com","Desk Lamp","Furniture",4,34.99,"=G29*F29","2024-02-06","South","Pending"],
    ["ORD-029","Claire Young","claire@example.com","USB Hub","Electronics",6,39.99,"=G30*F30","2024-02-07","North","Completed"],
    ["ORD-030","Derek King","derek@example.com","Standing Desk","Furniture",None,899.99,"=G31*F31","2024-02-08","East","Pending"],
    ["ORD-031","Ella Price","ella@example.com","Keyboard","Electronics",2,79.99,"=G32*F32","2024-02-09","West","Completed"],
    ["ORD-032","Felix Ward","felix@example.com","Webcam","Electronics",1,149.99,"=G33*F33","2024-02-10","South","Completed"],
    ["ORD-033","Gina Brooks","gina@example.com","Notebook","Stationery",100,4.99,"=G34*F34","2024-02-11","North","Completed"],
    ["ORD-034","Hank Cooper","hank@example.com","Wireless Mouse","Electronics",3,29.99,"=G35*F35","2024-02-12","East","Completed"],
    ["ORD-035","Isla Diaz","isla@example.com","Mouse Pad","Stationery",15,9.99,"=G36*F36","2024-02-13","West","Completed"],
    ["ORD-036","Jake Foster","jake@example.com","Laptop Pro","Electronics",1,1299.99,"=G37*F37","2024-02-14","South","Returned"],
    ["ORD-037","Kara Gomez","kara@example.com","Monitor 27in","Electronics",2,399.99,"=G38*F38","2024-02-15","North","Completed"],
    ["ORD-038","Liam Hayes","liam@example.com","Office Chair","Furniture",1,249.99,"=G39*F39","2024-02-16","East","Pending"],
    ["ORD-039","Mara Ingram","mara@example.com","Printer","Electronics",1,599.99,"=G40*F40","2024-02-17","West","Completed"],
    ["ORD-040","Neil James","neil@example.com","Desk Lamp","Furniture",8,34.99,"=G41*F41","2024-02-18","South","Completed"],
    ["ORD-041","Ora Klein","ora@example.com","USB Hub","Electronics",5,39.99,"=G42*F42","2024-02-19","North","Completed"],
    ["ORD-042","Pete Lawson","pete@example.com","Keyboard","Electronics",4,79.99,"=G43*F43","2024-02-20","East","Completed"],
    ["ORD-043","Quinn Ross","quinn@example.com","Standing Desk","Furniture",1,899.99,"=G44*F44","2024-02-21","West","Pending"],
    ["ORD-044","Rose Moore","","Wireless Mouse","Electronics",2,29.99,"=G45*F45","2024-02-22","South","Completed"],
    ["ORD-045","Sean Nash","sean@example.com","Notebook","Stationery",25,4.99,"=G46*F46","2024-02-23","North","Completed"],
    ["ORD-046","Tara Owens","tara@example.com","Mouse Pad","Stationery",10,9.99,"=G47*F47","2024-02-24","East","Completed"],
    ["ORD-047","Uma Flores","uma@example.com","Monitor 27in","Electronics",1,399.99,"=G48*F48","2024-02-25","West","Completed"],
    ["ORD-048","Val Peters","val@example.com","Webcam","Electronics",3,149.99,"=G49*F49","2024-02-26","South","Completed"],
    ["ORD-049","Will Quinn","will@example.com","Laptop Pro","Electronics",2,1299.99,"=G50*F50","2024-02-27","North","Returned"],
    ["ORD-050","Xena Rivera","xena@example.com","Printer","Electronics",1,599.99,"=G51*F51","2024-02-28","East","Completed"],
]

for row in sales_rows:
    ws1.append(row)

style_data_rows(ws1, headers1)

for i, w in enumerate([10,16,26,16,14,9,11,13,13,9,11], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"
wb1.save("D:/Projects/Promethean/data-probe/test/sales_data.xlsx")
print("sales_data.xlsx saved")

# ── employee_hr.xlsx ──────────────────────────────────────────────────────────
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Employees"

headers2 = ["employee_id","first_name","last_name","department","job_title",
            "salary","hire_date","email","phone","manager_id","is_active"]
ws2.append(headers2)
style_headers(ws2, len(headers2))

emp_rows = [
    ["E001","James","Carter","Engineering","Software Engineer",95000,"2021-03-15","james.carter@corp.com","555-0101","E010",True],
    ["E002","Sophia","Martin","HR","HR Manager",78000,"2019-07-01","sophia.martin@corp.com","555-0102","E015",True],
    ["E003","Liam","Thompson","Engineering","Senior Engineer",115000,"2018-05-20","liam.thompson@corp.com","555-0103","E010",True],
    ["E004","Olivia","White","Marketing","Marketing Lead",88000,"2020-11-30","olivia.white@corp.com","555-0104","E016",True],
    ["E005","Noah","Harris","Engineering","Junior Engineer",72000,"2022-01-10","noah.harris@corp.com","555-0105","E003",True],
    ["E006","Emma","Lewis","Finance","Accountant",82000,"2020-04-22","emma.lewis@corp.com","555-0106","E017",True],
    ["E007","Ava","Walker","Human Resources","HR Coordinator",61000,"2021-09-05","ava.walker@corp.com","555-0107","E002",True],
    ["E008","William","Hall","Engineering","DevOps Engineer",108000,"2019-02-14","william.hall@corp.com","555-0108","E010",True],
    ["E009","Isabella","Allen","Marketing","Content Writer",65000,"2023-06-01","isabella.allen@corp.com","555-0109","E004",True],
    ["E010","Mason","Young","Engineering","VP Engineering",145000,"2016-08-01","mason.young@corp.com","555-0110",None,True],
    ["E003","Liam","Thompson","Engineering","Senior Engineer",115000,"2018-05-20","liam.thompson@corp.com","555-0103","E010",True],
    ["E012","Mia","King","Finance","Financial Analyst",79000,"2022-03-18","mia.king@corp.com","555-0112","E017",True],
    ["E013","Ethan","Scott","Engineering","Software Engineer",96000,"2021-07-25","ethan.scott@corp.com","555-0113","E003",True],
    ["E014","Charlotte","Green","Sales","Sales Rep",70000,"2023-01-15","charlotte.green@corp.com","555-0114","E018",True],
    ["E015","Amelia","Baker","HR","Chief People Officer",132000,"2015-03-10","amelia.baker@corp.com","555-0115",None,True],
    ["E016","Benjamin","Adams","Marketing","VP Marketing",138000,"2017-06-20","benjamin.adams@corp.com","555-0116",None,True],
    ["E017","Harper","Nelson","Finance","CFO",160000,"2014-11-01","harper.nelson@corp.com","555-0117",None,True],
    ["E018","Elijah","Carter","Sales","VP Sales",142000,"2016-04-15","elijah.carter@corp.com","555-0118",None,True],
    ["E019","Abigail","Mitchell","Engineering","QA Engineer",85000,"2020-08-30","abigail.mitchell@corp.com","555-0119","E003",True],
    ["E020","Lucas","Perez","Sales","Sales Rep",68000,"2022-09-12","lucas.perez@corp.com","555-0120","E018",False],
    ["E021","Emily","Roberts","Finance","Accountant",-5000,"2021-12-01","emily.roberts@corp.com","555-0121","E017",True],
    ["E022","Henry","Turner","Marketing","SEO Specialist",67000,"2023-03-22","henry.turner@corp.com","555-0122","E004",True],
    ["E023","Evelyn","Phillips","Engineering","Software Engineer",94000,"2021-05-17","evelyn.phillips@corp.com","555-0123","E003",True],
    ["E024","Alexander","Campbell","HR","Recruiter",63000,"2022-07-08","alexander.campbell@corp.com","555-0124","E002",True],
    ["E025","Scarlett","Parker","Sales","Account Manager",75000,"2020-10-19","scarlett.parker@corp.com","555-0125","E018",True],
    ["E026","Daniel","Evans","Engineering","Data Engineer",101000,"2019-12-03","daniel.evans@corp.com","555-0126","E010",True],
    ["E027","Grace","Edwards","Finance","",74000,"2021-08-14","grace.edwards@corp.com","555-0127","E017",True],
    ["E028","Jackson","Collins","Mktg","Copywriter",62000,"2023-05-29","jackson.collins@corp.com","555-0128","E004",True],
    ["E029","Chloe","Stewart","Engineering","Software Engineer",None,"2026-02-01","chloe.stewart@corp.com","555-0129","E003",True],
    ["E030","Sebastian","Morris","Sales","Sales Rep",71000,"2022-11-07","","555-0130","E018",False],
]

for row in emp_rows:
    ws2.append(row)

style_data_rows(ws2, headers2)

for i, w in enumerate([13,12,13,16,20,10,12,28,13,12,10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A2"
wb2.save("D:/Projects/Promethean/data-probe/test/employee_hr.xlsx")
print("employee_hr.xlsx saved")
