"""
Generates clean_orders.xlsx — a deliberately perfect dataset for scoring tests.

Design goals (all must hold to achieve health_score == 100):
  - Zero NULL values in every column
  - All date values in strict ISO-8601 format (YYYY-MM-DD)
  - All email values in valid format
  - All numeric columns contain only numeric strings (no mixed types)
  - total_price is always quantity * unit_price (no formula drift)
  - Enum columns (category, status, region, payment_method) use consistent,
    identical casing with no variants
  - order_id is unique across all rows (near-unique cardinality flag only)
  - No duplicate rows
  - No outlier values (prices in a realistic range, quantities > 0)
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="1B4332")
CELL_FONT = Font(name="Arial", size=10)


def style_sheet(ws, col_count: int, col_widths: list[int]) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in range(2, ws.max_row + 1):
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).font = CELL_FONT
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


PRODUCTS = [
    ("SKU-001", "Laptop Pro",      "Electronics", 1299.99),
    ("SKU-002", "Office Chair",    "Furniture",    249.99),
    ("SKU-003", "Wireless Mouse",  "Electronics",   29.99),
    ("SKU-004", "Standing Desk",   "Furniture",    899.99),
    ("SKU-005", "Monitor 27in",    "Electronics",  399.99),
    ("SKU-006", "Keyboard",        "Electronics",   79.99),
    ("SKU-007", "Webcam",          "Electronics",  149.99),
    ("SKU-008", "Notebook",        "Stationery",     4.99),
    ("SKU-009", "USB Hub",         "Electronics",   39.99),
    ("SKU-010", "Desk Lamp",       "Furniture",     34.99),
    ("SKU-011", "Mouse Pad",       "Stationery",     9.99),
    ("SKU-012", "Printer",         "Electronics",  599.99),
]

CUSTOMERS = [
    ("CUST-01", "Alice Johnson",   "alice.johnson@example.com"),
    ("CUST-02", "Bob Smith",       "bob.smith@example.com"),
    ("CUST-03", "Carol White",     "carol.white@example.com"),
    ("CUST-04", "David Lee",       "david.lee@example.com"),
    ("CUST-05", "Eve Davis",       "eve.davis@example.com"),
    ("CUST-06", "Frank Brown",     "frank.brown@example.com"),
    ("CUST-07", "Grace Kim",       "grace.kim@example.com"),
    ("CUST-08", "Henry Clark",     "henry.clark@example.com"),
    ("CUST-09", "Iris Lopez",      "iris.lopez@example.com"),
    ("CUST-10", "Jack Patel",      "jack.patel@example.com"),
]

REGIONS = ["North", "South", "East", "West"]
STATUSES = ["Completed", "Pending", "Returned"]
PAYMENTS = ["Credit Card", "Bank Transfer", "PayPal"]

# 60 rows: every combination of product/customer cycles cleanly
ROWS = [
    # order_id, cust_idx, prod_idx, qty, date, status, region, payment
    ( 1,  0,  0, 1, "2024-01-05", "Completed", "North", "Credit Card"),
    ( 2,  1,  1, 4, "2024-01-07", "Completed", "South", "Bank Transfer"),
    ( 3,  2,  2, 10,"2024-01-09", "Completed", "East",  "PayPal"),
    ( 4,  3,  3, 1, "2024-01-11", "Pending",   "West",  "Credit Card"),
    ( 5,  4,  4, 3, "2024-01-12", "Completed", "North", "Bank Transfer"),
    ( 6,  5,  5, 5, "2024-01-14", "Completed", "East",  "PayPal"),
    ( 7,  6,  6, 2, "2024-01-15", "Returned",  "South", "Credit Card"),
    ( 8,  7,  7, 50,"2024-01-16", "Completed", "West",  "Bank Transfer"),
    ( 9,  8,  0, 1, "2024-01-18", "Completed", "North", "Credit Card"),
    (10,  9,  8, 8, "2024-01-20", "Completed", "East",  "PayPal"),
    (11,  1,  1, 2, "2024-01-21", "Completed", "South", "Credit Card"),
    (12,  0,  11,1, "2024-01-22", "Pending",   "North", "Bank Transfer"),
    (13,  2,  9, 6, "2024-01-23", "Completed", "West",  "PayPal"),
    (14,  3,  6, 1, "2024-01-24", "Completed", "East",  "Credit Card"),
    (15,  4,  0, 1, "2024-01-25", "Completed", "South", "Bank Transfer"),
    (16,  5,  10,20,"2024-01-26", "Completed", "North", "PayPal"),
    (17,  6,  4, 2, "2024-01-27", "Pending",   "East",  "Credit Card"),
    (18,  7,  3, 1, "2024-01-28", "Completed", "West",  "Bank Transfer"),
    (19,  8,  5, 3, "2024-01-29", "Completed", "South", "PayPal"),
    (20,  9,  8, 4, "2024-01-30", "Completed", "North", "Credit Card"),
    (21,  0,  7, 30,"2024-01-31", "Completed", "East",  "Bank Transfer"),
    (22,  1,  0, 2, "2024-02-01", "Completed", "West",  "PayPal"),
    (23,  2,  1, 3, "2024-02-02", "Completed", "South", "Credit Card"),
    (24,  3,  2, 7, "2024-02-03", "Returned",  "North", "Bank Transfer"),
    (25,  4,  4, 1, "2024-02-04", "Completed", "East",  "PayPal"),
    (26,  5,  11,2, "2024-02-05", "Completed", "West",  "Credit Card"),
    (27,  6,  9, 5, "2024-02-06", "Pending",   "South", "Bank Transfer"),
    (28,  7,  9, 4, "2024-02-07", "Completed", "North", "PayPal"),
    (29,  8,  3, 1, "2024-02-08", "Pending",   "East",  "Credit Card"),
    (30,  9,  5, 2, "2024-02-09", "Completed", "West",  "Bank Transfer"),
    (31,  0,  6, 1, "2024-02-10", "Completed", "South", "PayPal"),
    (32,  1,  7, 100,"2024-02-11","Completed", "North", "Credit Card"),
    (33,  2,  2, 3, "2024-02-12", "Completed", "East",  "Bank Transfer"),
    (34,  3,  10,15,"2024-02-13", "Completed", "West",  "PayPal"),
    (35,  4,  0, 1, "2024-02-14", "Returned",  "South", "Credit Card"),
    (36,  5,  4, 2, "2024-02-15", "Completed", "North", "Bank Transfer"),
    (37,  6,  1, 1, "2024-02-16", "Pending",   "East",  "PayPal"),
    (38,  7,  11,1, "2024-02-17", "Completed", "West",  "Credit Card"),
    (39,  8,  9, 8, "2024-02-18", "Completed", "South", "Bank Transfer"),
    (40,  9,  8, 5, "2024-02-19", "Completed", "North", "PayPal"),
    (41,  0,  5, 4, "2024-02-20", "Completed", "East",  "Credit Card"),
    (42,  1,  3, 1, "2024-02-21", "Pending",   "West",  "Bank Transfer"),
    (43,  2,  2, 2, "2024-02-22", "Completed", "South", "PayPal"),
    (44,  3,  7, 25,"2024-02-23", "Completed", "North", "Credit Card"),
    (45,  4,  10,10,"2024-02-24", "Completed", "East",  "Bank Transfer"),
    (46,  5,  4, 1, "2024-02-25", "Completed", "West",  "PayPal"),
    (47,  6,  6, 3, "2024-02-26", "Completed", "South", "Credit Card"),
    (48,  7,  0, 2, "2024-02-27", "Returned",  "North", "Bank Transfer"),
    (49,  8,  11,1, "2024-02-28", "Completed", "East",  "PayPal"),
    (50,  9,  1, 1, "2024-03-01", "Completed", "West",  "Credit Card"),
    (51,  0,  8, 6, "2024-03-02", "Completed", "North", "Bank Transfer"),
    (52,  1,  5, 2, "2024-03-03", "Completed", "South", "PayPal"),
    (53,  2,  3, 1, "2024-03-04", "Pending",   "East",  "Credit Card"),
    (54,  3,  0, 3, "2024-03-05", "Completed", "West",  "Bank Transfer"),
    (55,  4,  7, 40,"2024-03-06", "Completed", "North", "PayPal"),
    (56,  5,  2, 8, "2024-03-07", "Completed", "South", "Credit Card"),
    (57,  6,  4, 1, "2024-03-08", "Completed", "East",  "Bank Transfer"),
    (58,  7,  10,12,"2024-03-09", "Completed", "West",  "PayPal"),
    (59,  8,  6, 2, "2024-03-10", "Completed", "North", "Credit Card"),
    (60,  9,  11,1, "2024-03-11", "Completed", "South", "Bank Transfer"),
]

headers = [
    "order_id", "customer_id", "customer_name", "customer_email",
    "product_sku", "product_name", "category",
    "quantity", "unit_price", "total_price",
    "order_date", "status", "region", "payment_method",
]

wb = Workbook()
ws = wb.active
ws.title = "Orders"
ws.append(headers)

for (num, ci, pi, qty, date, status, region, payment) in ROWS:
    sku, name, category, price = PRODUCTS[pi]
    cust_id, cust_name, email = CUSTOMERS[ci]
    total = round(qty * price, 2)
    ws.append([
        f"ORD-{num:04d}",
        cust_id,
        cust_name,
        email,
        sku,
        name,
        category,
        qty,
        price,
        total,
        date,
        status,
        region,
        payment,
    ])

col_widths = [10, 10, 16, 28, 10, 16, 14, 10, 11, 12, 12, 11, 8, 14]
style_sheet(ws, len(headers), col_widths)

out_path = "clean_orders.xlsx"
wb.save(out_path)
print(f"Saved {out_path} ({len(ROWS)} rows, {len(headers)} columns)")
print()
print("Scoring properties:")
print("  - Zero NULL values")
print("  - All dates in ISO-8601 format (YYYY-MM-DD)")
print("  - All emails in valid format")
print("  - All numeric columns contain only numeric values")
print("  - total_price = quantity * unit_price (no formula drift)")
print("  - Enum columns use consistent casing (no variants)")
print("  - order_id is unique (60 distinct values, near-unique flag only)")
print("  - No duplicate rows")
print("  - No outlier values")
print("  => Expected health_score: 100")
